"""Parse uploaded governance Excel files."""

from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.excel.template_generator import TEMPLATE_HEADERS

REQUIRED_COLUMNS = {"metric_code", "value"}


@dataclass
class ParsedExcelRow:
    row_number: int
    metric_code: str
    raw_value: str | None
    metric_name: str | None = None
    dimension: str | None = None


class ExcelParseError(Exception):
    def __init__(self, message: str) -> None:
        self.message = message
        super().__init__(message)


class ExcelParserService:
    """Read xlsx and extract metric rows (does not validate business rules)."""

    def parse(self, file_bytes: bytes) -> list[ParsedExcelRow]:
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelParseError("File is not a valid Excel workbook") from exc

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ExcelParseError("Workbook is empty") from exc

        headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
        if not headers or not any(headers):
            raise ExcelParseError("Missing header row")

        col_index = {name: idx for idx, name in enumerate(headers) if name}
        missing = REQUIRED_COLUMNS - set(col_index)
        if missing:
            raise ExcelParseError(f"Missing required columns: {', '.join(sorted(missing))}")

        code_idx = col_index["metric_code"]
        value_idx = col_index["value"]
        name_idx = col_index.get("metric_name")
        dim_idx = col_index.get("dimension")

        parsed: list[ParsedExcelRow] = []
        row_number = 1
        for excel_row in rows_iter:
            row_number += 1
            if excel_row is None or all(cell is None or str(cell).strip() == "" for cell in excel_row):
                continue

            code_cell = excel_row[code_idx] if code_idx < len(excel_row) else None
            metric_code = str(code_cell).strip() if code_cell is not None else ""
            if not metric_code:
                continue

            value_cell = excel_row[value_idx] if value_idx < len(excel_row) else None
            raw_value = None if value_cell is None else str(value_cell).strip()

            metric_name = None
            if name_idx is not None and name_idx < len(excel_row) and excel_row[name_idx] is not None:
                metric_name = str(excel_row[name_idx]).strip()

            dimension = None
            if dim_idx is not None and dim_idx < len(excel_row) and excel_row[dim_idx] is not None:
                dimension = str(excel_row[dim_idx]).strip()

            parsed.append(
                ParsedExcelRow(
                    row_number=row_number,
                    metric_code=metric_code,
                    raw_value=raw_value if raw_value != "" else None,
                    metric_name=metric_name,
                    dimension=dimension,
                )
            )

        wb.close()
        if not parsed:
            raise ExcelParseError("No metric rows found in file")
        return parsed

    @staticmethod
    def expected_headers() -> list[str]:
        return list(TEMPLATE_HEADERS)
